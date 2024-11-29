import os, sys
# import traceback
if hasattr(sys, "frozen"):
    os.environ["PATH"] = sys._MEIPASS + ";" + os.environ["PATH"]
from PyQt5 import QtWidgets, QtCore
from openpyxl import load_workbook
import xlwt
from collections import defaultdict


class ExcelSplitterApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel拆分工具")
        self.setGeometry(200, 200, 600, 600)
        self.setAcceptDrops(True)

        # 初始化界面
        self.file_path = ""
        self.sheet_list = []
        self.selected_sheets = []

        self.init_ui()

    def init_ui(self):
        # 文件选择区
        self.file_label = QtWidgets.QLabel("拖入Excel文件或点击“浏览”按钮选择文件。")
        self.file_label.setStyleSheet(
            "background-color: #F7F8FA; border: 1px solid #ccc; padding: 10px;"
        )
        self.file_label.setAlignment(QtCore.Qt.AlignCenter)
        self.file_label.setFixedHeight(100)

        self.browse_button = QtWidgets.QPushButton("浏览")
        self.browse_button.clicked.connect(self.browse_file)

        # Sheet 选择区
        self.sheet_list_widget = QtWidgets.QListWidget()
        self.sheet_list_widget.setSelectionMode(
            QtWidgets.QAbstractItemView.MultiSelection
        )

        # 进度条
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setTextVisible(False)
        # self.progress_bar.setValue(0)

        # 操作按钮
        self.process_button = QtWidgets.QPushButton("处理并导出")
        self.process_button.clicked.connect(self.process_and_export)

        # 布局
        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.file_label)
        layout.addWidget(self.browse_button)
        layout.addWidget(QtWidgets.QLabel("选择需要处理的Sheet："))
        layout.addWidget(self.sheet_list_widget)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.process_button)
        self.setLayout(layout)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.endswith((".xlsx", ".xls")):
                self.load_file(file_path)
            else:
                QtWidgets.QMessageBox.warning(
                    self, "无效文件", "请拖入有效的Excel文件！"
                )

    def browse_file(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        if file_path:
            self.load_file(file_path)

    def load_file(self, file_path):
        self.file_path = file_path
        self.file_label.setText(f"已加载文件：{os.path.basename(file_path)}")
        try:
            workbook = load_workbook(file_path, read_only=True)
            self.sheet_list = workbook.sheetnames

            # 更新 Sheet 列表并显示行数
            self.sheet_list_widget.clear()
            for sheet_name in self.sheet_list:
                sheet = workbook[sheet_name]
                row_count = sheet.max_row
                self.sheet_list_widget.addItem(f"{sheet_name}（{row_count}行）")
            if self.sheet_list:
                self.sheet_list_widget.item(0).setSelected(True)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "加载失败", f"无法加载文件：{e}")

    def process_and_export(self):
        if not self.file_path:
            QtWidgets.QMessageBox.warning(self, "未选择文件", "请先加载一个Excel文件！")
            return

        # 获取选中的 Sheets
        selected_items = self.sheet_list_widget.selectedItems()
        self.selected_sheets = [item.text().split("（")[0] for item in selected_items]

        if not self.selected_sheets:
            QtWidgets.QMessageBox.warning(self, "未选择Sheet", "请至少选择一个Sheet！")
            return

        try:
            self.merge_and_export()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "处理失败", f"处理过程中出错：{e}")

    def is_not_blank(self, item):
        if str(item).strip() == "--":
            return False
        elif len(str(item).strip()) == 0:
            return False
        elif item and str(item).strip():
            return True
        else:
            return False

    def merge_and_export(self):
        """从多个 sheet 合并数据（去重并合并相同发票号码或数电票号码的数据），并导出"""
        try:
            # 打开文件
            workbook = load_workbook(self.file_path)
            
            # 存储每个发票号码或数电票号码对应的合并数据
            invoice_map = defaultdict(
                lambda: {
                    "发票类型": "",
                    "发票代码": "",
                    "发票号码": "",
                    "数电票号码": "",
                    "开票日期": "",
                    "价税合计": 0.0,
                    "校验码": "",
                }
            )

            # 遍历选择的 sheet
            for sheet_name in self.selected_sheets:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                # 如果没有内容跳过
                if len(rows) < 2:
                    continue

                # 获取标题行和数据行
                headers = rows[0]
                data_rows = rows[1:]

                header_index = {header: idx for idx, header in enumerate(headers)}

                # 按输入表字段收集数据
                for row in data_rows:
                    invoice_type = str(row[header_index["发票票种"]]).strip() if row[header_index["发票票种"]] else None
                    invoice_number = str(row[header_index["发票号码"]]).strip() if row[header_index["发票号码"]] else None
                    electronic_invoice_number = str(row[header_index["数电票号码"]]).strip() if row[header_index["数电票号码"]] else None
                    invoice_code = str(row[header_index["发票代码"]]).strip() if row[header_index["发票代码"]] else ""
                    invoice_date = str(row[header_index["开票日期"]]).split()[0] if row[header_index["开票日期"]] else ""
                    tax_and_total = float(row[header_index["价税合计"]]) if row[header_index["价税合计"]] else 0.0

                    # 合并“发票号码”或“数电票号码”相同的行
                    if invoice_number:
                        invoice_key = invoice_number
                    elif electronic_invoice_number:
                        invoice_key = electronic_invoice_number
                    else:
                        continue  # 如果两者都为空则跳过

                    # 合并数据并求和
                    invoice_map[invoice_key]['发票票种'] = invoice_type if invoice_type else ""
                    invoice_map[invoice_key]['发票代码'] = invoice_code if invoice_code else ""
                    invoice_map[invoice_key]['发票号码'] = invoice_number if invoice_number else electronic_invoice_number
                    invoice_map[invoice_key]['数电票号码'] = electronic_invoice_number if electronic_invoice_number else ""
                    invoice_map[invoice_key]['开票日期'] = invoice_date  # 保证日期一致
                    invoice_map[invoice_key]['价税合计'] += tax_and_total  # 累加税价合计
                    invoice_map[invoice_key]['校验码'] = ""  # 校验码固定为空

            total_rows = len(invoice_map)
            self.progress_bar.setMaximum(total_rows)

            # 按需处理导出的字段
            output_data = []
            for invoice_key, data in invoice_map.items():
                mapped_row = []
                # 纸质发票-普通发票
                if self.is_not_blank(data["发票号码"]) and not self.is_not_blank(
                    data["发票代码"]
                ):
                    mapped_row = [
                        # "增值税普通发票",  # 发票类型
                        str(data["发票票种"]).strip() if data["发票票种"] else "",
                        str(data["发票代码"]).strip() if data["发票代码"] else "",  # 发票代码
                        str(data["发票号码"]).strip() if data["发票号码"] else "",  # 发票号码
                        str(data["开票日期"]).split()[0] if data["开票日期"] else "",  # 开票日期
                        str(data["价税合计"]).strip() if data["价税合计"] else "",  # 价税合计
                        "",  # 校验码，固定为空
                    ]
                # 纸质发票-专用发票
                if self.is_not_blank(data["发票号码"]) and self.is_not_blank(
                    data["发票代码"]
                ):
                    mapped_row = [
                        # "增值税专用发票",  
                        str(data["发票票种"]).strip() if data["发票票种"] else "", # 发票类型
                        str(data["发票代码"]).strip() if data["发票代码"] else "",  # 发票代码
                        str(data["发票号码"]).strip() if data["发票号码"] else "",  # 发票号码
                        str(data["开票日期"]).split()[0] if data["开票日期"] else "",  # 开票日期
                        str(data["价税合计"]).strip() if data["价税合计"] else "",  # 价税合计
                        "",  # 校验码，固定为空
                    ]
                # 电子发票
                if self.is_not_blank(data["数电票号码"]):
                    mapped_row = [
                        "数电发票（专票）",  # 发票类型
                        "",  # 发票代码
                        str(data["数电票号码"]).strip() if data["数电票号码"] else "",  # 发票号码
                        str(data["开票日期"]).split()[0] if data["开票日期"] else "",  # 开票日期
                        str(data["价税合计"]).strip() if data["价税合计"] else "",  # 价税合计
                        "",  # 校验码，固定为空
                    ]

                mapped_row = [
                    str(item).strip() if item is not None else "" for item in mapped_row
                ]
                output_data.append(mapped_row)

            # 导出
            output_dir = QtWidgets.QFileDialog.getExistingDirectory(
                self, "选择导出目录"
            )
            if not output_dir:
                return

            chunk_size = 20
            file_count = 0
            for i in range(0, len(output_data), chunk_size):
                chunk = output_data[i : i + chunk_size]
                output_path = os.path.join(output_dir, f"fp{file_count + 1}.xls")
                self.progress_bar.setValue(i + len(chunk))
                self.export_to_excel(chunk, output_path)
                file_count += 1

            # 提示用户完成
            QtWidgets.QMessageBox.information(
                self,
                "拆分成功",
                f"拆分成功：\n总数据 {len(output_data)} 条\n拆分文件 {file_count} 个。\n"
                f"拆分后的文件保存于：{output_dir}"
                f"\n\n注意：增值税电子普通发票、增值税普通发票这两种类型，需在pt开头文件中补充校验码后6位。",
            )

        except Exception as e:
            # 调试使用
            # error_msg = traceback.format_exc()  # 获取完整的堆栈信息
            # QtWidgets.QMessageBox.critical(self, "错误", f"处理出错：\n{error_msg}")
            QtWidgets.QMessageBox.critical(self, "错误", f"处理出错：{str(e)}")

    def export_to_excel(self, data, file_path):
        """导出数据到 Excel 文件"""
        # 创建工作簿和工作表
        workbook = xlwt.Workbook(encoding="utf-8")
        sheet = workbook.add_sheet("拆分结果", cell_overwrite_ok=True)

        # 设置表头
        headers = ["发票类型", "发票代码", "发票号码", "开票日期", "金额", "校验码"]
        for col, header in enumerate(headers):
            sheet.write(0, col, header)

        # 写入数据
        for row_num, row in enumerate(data, start=1):
            for col_num, value in enumerate(row):
                sheet.write(row_num, col_num, value)

        # 设置文本格式（左对齐）
        style = xlwt.XFStyle()
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_LEFT
        style.alignment = alignment

        for row_num in range(1, len(data) + 1):
            for col_num in range(len(headers)):
                sheet.write(row_num, col_num, data[row_num - 1][col_num], style)

        # 保存文件
        workbook.save(file_path)


def load_stylesheet():
    return """
QPushButton {
  background-color: #5478ba;
  color: white;
  border: none;
  padding: 4px 4px;
  border-radius: 5px;
}

QPushButton:pressed {
  background-color: #436094; /* 鼠标按下时的颜色 */
}

QPushButton:hover {
  background-color: #436094; /* 鼠标按下时的颜色 */
}

QListWidget {
  background-color: #FFFFFF; /* 背景 */
  border: 1px solid #ccc;
}

QProgressBar {
  color: #101825; 
  height: 1px;
  border: 1px solid #ccc;
  background-color: #FFFFFF;
}

QProgressBar::chunk {
  background-color: #06C687;   /* 进度部分的颜色 */
  width: 3px;
}

QListWidget::item:selected {
  background-color: #436094;  /* 选中项背景色 */
  color: #ffffff;             /* 选中项文本颜色 */
}
"""


if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    app.setStyleSheet(load_stylesheet())

    window = ExcelSplitterApp()
    window.show()
    app.exec_()
